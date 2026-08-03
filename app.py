import streamlit as st
import pandas as pd
import json
import re
import requests
import os
from io import BytesIO
from copy import deepcopy
from dotenv import load_dotenv

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font

# ============================================================
#  APIキー読み込み
# ============================================================
load_dotenv()
API_KEY = os.getenv("GEMINI_API_KEY")

if not API_KEY:
    st.error("❌ GEMINI_API_KEY が読み込めていません。.env を確認してください。")
    st.stop()

# ============================================================
#  Gemini REST API（3.6 Flash）
# ============================================================
def gemini_generate(prompt: str) -> str:
    url = (
        "https://generativelanguage.googleapis.com/v1beta/"
        "models/gemini-3.6-flash:generateContent?key=" + API_KEY
    )
    headers = {"Content-Type": "application/json"}
    data = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0, "maxOutputTokens": 8192},
    }

    # 503 対策：最大3回まで自動リトライ
    for attempt in range(3):
        response = requests.post(url, headers=headers, json=data)
        result = response.json()

        if "error" not in result:
            return result["candidates"][0]["content"]["parts"][0]["text"]

        # 503 の場合は再試行
        if result["error"].get("code") == 503:
            time.sleep(2)  # 少し待つ
            continue

        # その他のエラーは即停止
        st.error("Gemini API エラー:")
        st.json(result)
        st.stop()

    st.error("Gemini API が混雑のため応答できませんでした（503）。時間を置いて再試行してください。")
    st.stop()

# ============================================================
#  勤務記号（文化OS ver3.0）
# ============================================================
CODE_LIST = [
    "公", "明公",
    "早1", "早A", "早B", "早C",
    "日1", "日2ホ",
    "遅1A", "遅1B", "遅1C",
    "夜1", "夜2"
]
# ============================================================
#  勤務記号入りシート自動判定
# ============================================================
def find_sheet_with_codes(dfs: dict) -> str | None:
    pattern = "|".join(CODE_LIST)
    for name, df in dfs.items():
        try:
            if df.apply(lambda row: row.astype(str).str.contains(pattern).any(), axis=1).any():
                return name
        except Exception:
            continue
    return None

# ============================================================
#  職員名判定（漢字・ひらがな・カタカナのみ）
# ============================================================
def is_staff_name(text: str) -> bool:
    if not text:
        return False
    t = text.replace("☆", "").strip()

    # 除外ワード
    if t in ["月間予定"]:
        return False
    if any(x in t for x in ["～", ":", "勤務時間", "週", "月～金", "土日祝"]):
        return False
    if any(x in t for x in ["グループ", "介護長", "介護主任", "新入職員", "パート"]):
        return False

    # 漢字・ひらがな・カタカナのみ
    if re.match(r'^[\u4E00-\u9FFF\u3040-\u309F\u30A0-\u30FF]+$', t):
        return True
    return False

# ============================================================
#  勤務記号抽出（確実に拾える安定版）
# ============================================================
def extract_codes(df_raw: pd.DataFrame):
    codes = set()

    for _, row in df_raw.iterrows():
        for cell in row:
            text = str(cell).strip()

            # 空欄はスキップ
            if not text or text == "nan":
                continue

            # 勤務記号リストのどれかが含まれていれば追加
            for code in CODE_LIST:
                if code in text:
                    codes.add(code)

            # 数字＋記号の複合表記にも対応（例：早1A、遅1B、日2ホ）
            for code in CODE_LIST:
                if re.search(rf"{code}", text):
                    codes.add(code)

            # 「早」「日」「遅」「夜」だけの表記にも対応
            if any(k in text for k in ["早", "日", "遅", "夜"]):
                for code in CODE_LIST:
                    if code.startswith(text[0]):  # 例：早 → 早1, 早A, 早B...
                        codes.add(code)

    return sorted(list(codes))

# ============================================================
#  役職・グループ抽出（まーくん勤務表構造完全対応版）
# ============================================================
def extract_roles_groups(df_raw: pd.DataFrame, filtered_names):
    roles = {}
    groups = {}

    current_role = None
    current_group = None

    for _, row in df_raw.iterrows():
        row_values = [str(x).strip() for x in row.tolist()]

        # 役職行
        if "介護長" in row_values:
            current_role = "介護長"
            current_group = None
            continue

        if "介護主任" in row_values:
            current_role = "介護主任"
            current_group = None
            continue

        if "新入職員" in row_values:
            current_role = "新入職員"
            current_group = None
            continue

        if "パート" in row_values:
            current_role = "パート"
            current_group = None
            continue

        # グループ行
        if "Aグループ" in row_values:
            current_group = "Aグループ"
            current_role = None
            continue

        if "Bグループ" in row_values:
            current_group = "Bグループ"
            current_role = None
            continue

        if "Cグループ" in row_values:
            current_group = "Cグループ"
            current_role = None
            continue

        # 職員名行
        for name in filtered_names:
            if name in row_values:
                if current_role:
                    roles[name] = current_role
                if current_group:
                    groups[name] = current_group

    return roles, groups

# ============================================================
#  既存勤務表の書式抽出（色・罫線・セル結合・列幅・行高さ）
# ============================================================
def extract_format_from_existing_excel(uploaded_file, sheet_name):
    # openpyxl で読み直す
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file)
    ws = wb[sheet_name]

    format_map = {}

    # セルごとの書式を全部吸い取る
    for row in ws.iter_rows():
        for cell in row:
            format_map[(cell.row, cell.column)] = {
                "fill": cell.fill,
                "border": cell.border,
                "font": cell.font,
                "alignment": cell.alignment,
            }

    # 列幅
    col_widths = {
        col: ws.column_dimensions[col].width
        for col in ws.column_dimensions
    }

    # 行高さ
    row_heights = {
        row: ws.row_dimensions[row].height
        for row in ws.row_dimensions
    }

    # セル結合
    merged_cells = list(ws.merged_cells.ranges)

    return format_map, col_widths, row_heights, merged_cells
# ============================================================
#  生成勤務表に既存書式を適用（完全コピー）
# ============================================================
from openpyxl.styles import NamedStyle

def apply_format_to_generated_sheet(generated_data, format_map, col_widths, row_heights, merged_cells):
    wb = Workbook()
    ws = wb.active
    ws.title = "勤務表"

    # セル結合
    for mc in merged_cells:
        ws.merge_cells(str(mc))

    # 列幅
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # 行高さ
    for row, height in row_heights.items():
        ws.row_dimensions[row].height = height

    # 書式コピー用 NamedStyle キャッシュ
    style_cache = {}

    for r_idx, row in enumerate(generated_data, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if (r_idx, c_idx) in format_map:
                fmt = format_map[(r_idx, c_idx)]

                # キャッシュキー
                key = (fmt["fill"], fmt["border"], fmt["font"], fmt["alignment"])

                if key not in style_cache:
                    style = NamedStyle(name=f"style_{r_idx}_{c_idx}")
                    style.fill = fmt["fill"]
                    style.border = fmt["border"]
                    style.font = fmt["font"]
                    style.alignment = fmt["alignment"]
                    wb.add_named_style(style)
                    style_cache[key] = style

                cell.style = style_cache[key]

    return wb

# ============================================================
#  Streamlit UI（ページ設定）
# ============================================================
st.set_page_config(page_title="勤務表AI（完全統合版）", layout="wide")
st.title("📘 勤務表AI（Gemini 3.6 Flash / 本番仕様＋完全コピー）")

# ============================================================
#  Excel アップロード
# ============================================================
uploaded_file = st.sidebar.file_uploader("勤務表Excelをアップロード", type=["xlsx", "xlsm"])

if not uploaded_file:
    st.info("Excel をアップロードしてください。")
    st.stop()

# ============================================================
#  Excel 全シート読み込み
# ============================================================
xls = pd.ExcelFile(uploaded_file)
sheets = xls.sheet_names
dfs = {name: pd.read_excel(uploaded_file, sheet_name=name, header=None) for name in sheets}

st.write("### 読み込んだシート一覧")
st.json(sheets)

# ============================================================
#  勤務記号入りシートを自動判定
# ============================================================
target_sheet = find_sheet_with_codes(dfs)
if not target_sheet:
    st.error("❌ 勤務記号が含まれるシートが見つかりませんでした。")
    st.stop()

st.success(f"勤務記号入りシートを検出しました： {target_sheet}")

df_raw = dfs[target_sheet]

# ============================================================
#  職員名抽出
# ============================================================
name_col = df_raw.iloc[:, 1].fillna("").astype(str).tolist()
filtered_names = [n.replace("☆", "") for n in name_col if is_staff_name(n)]

st.write("### 抽出された職員名")
st.json(filtered_names)

# ============================================================
#  勤務記号抽出
# ============================================================
code_candidates = extract_codes(df_raw)

st.write("### 抽出された勤務記号（候補）")
st.json(code_candidates)

# ============================================================
#  役職・グループ抽出（まーくん勤務表構造）
# ============================================================
roles, groups = extract_roles_groups(df_raw, filtered_names)

st.write("### 役職候補")
st.json(roles)

st.write("### グループ候補")
st.json(groups)

# ============================================================
#  既存勤務表の書式抽出（色・罫線・セル結合・列幅・行高さ）
# ============================================================
format_map, col_widths, row_heights, merged_cells = extract_format_from_existing_excel(uploaded_file, target_sheet)
# ============================================================
# 解析関数セット（Part7 の直前に必ず置く）
# ============================================================

import re
import json
import pandas as pd

# ----------------------------------------
# ① 職員名抽出（漢字2〜4文字を名前と判定）
# ----------------------------------------
def extract_names(df_raw):
    names = set()
    for _, row in df_raw.iterrows():
        for cell in row:
            text = str(cell).strip()
            # まーくん勤務表の名前は漢字2〜4文字
            if re.match(r'^[\u4E00-\u9FFF]{2,4}$', text):
                names.add(text)
    return sorted(list(names))


# ----------------------------------------
# ② 役職・グループ抽出（まーくん勤務表構造）
# ----------------------------------------
def extract_roles_groups(df_raw, filtered_names):
    roles = {}
    groups = {}

    current_role = None
    current_group = None

    for _, row in df_raw.iterrows():
        row_values = [str(x).strip() for x in row.tolist()]

        # 役職行
        if "介護長" in row_values:
            current_role = "介護長"; current_group = None; continue
        if "介護主任" in row_values:
            current_role = "介護主任"; current_group = None; continue
        if "新入職員" in row_values:
            current_role = "新入職員"; current_group = None; continue
        if "パート" in row_values:
            current_role = "パート"; current_group = None; continue

        # グループ行
        if "Aグループ" in row_values:
            current_group = "Aグループ"; current_role = None; continue
        if "Bグループ" in row_values:
            current_group = "Bグループ"; current_role = None; continue
        if "Cグループ" in row_values:
            current_group = "Cグループ"; current_role = None; continue

        # 職員名行
        for name in filtered_names:
            if name in row_values:
                if current_role:
                    roles[name] = current_role
                if current_group:
                    groups[name] = current_group

    return roles, groups


# ----------------------------------------
# ③ 既存勤務表の読み取り（新人・パート固定用）
# ----------------------------------------
def extract_existing_schedule(df_raw, filtered_names):
    schedules = {name: {} for name in filtered_names}

    for _, row in df_raw.iterrows():
        row_values = [str(x).strip() for x in row.tolist()]

        for name in filtered_names:
            if name in row_values:
                idx = row_values.index(name)

                # 名前の右側に 1〜30日の勤務記号が並んでいる前提
                for day in range(1, 31):
                    if idx + day < len(row_values):
                        schedules[name][str(day)] = row_values[idx + day]

    return schedules


# ----------------------------------------
# ④ 勤務記号抽出（CODE_LIST に含まれる記号を拾う）
# ----------------------------------------
def extract_codes(df_raw):
    codes = set()
    for _, row in df_raw.iterrows():
        for cell in row:
            text = str(cell).strip()
            for code in CODE_LIST:
                if code and code in text:
                    codes.add(code)
    return sorted(list(codes))

# ============================================================
# ① 既存勤務表を解析する（職員・役職・グループ・勤務記号）
# ============================================================
st.write("## ① 既存勤務表を解析する")

uploaded_file = st.file_uploader("既存勤務表（Excel）をアップロード", type=["xlsx", "xlsm"])

if uploaded_file:
    # xlsm対応（openpyxlで読み取る）
    df_raw = pd.read_excel(uploaded_file, header=None, engine="openpyxl")

    # 職員名抽出
    filtered_names = extract_names(df_raw)

    # 役職・グループ抽出（強化版）
    roles, groups = extract_roles_groups(df_raw, filtered_names)

    # 勤務記号抽出
    code_candidates = extract_codes(df_raw)

    # 職員ごとの既存勤務表（新人・パート固定用）
    staff_schedules = extract_existing_schedule(df_raw, filtered_names)

    # JSON化
    parsed_staff = []
    for name in filtered_names:
        parsed_staff.append({
            "name": name,
            "role": roles.get(name, ""),
            "group": groups.get(name, ""),
            "schedule": staff_schedules.get(name, {})
        })

    st.session_state["parsed_staff"] = parsed_staff
    st.session_state["parsed_codes"] = code_candidates

    st.success("既存勤務表を解析しました！（xlsm対応）")
    st.json(parsed_staff)


# ============================================================
# ② 翌月の勤務表を生成する（新人・パート固定）
# ============================================================
st.write("## ② 翌月の勤務表を生成する（本番仕様JSON）")

default_month = "2026-06"
month = st.text_input("生成する月（YYYY-MM形式）", value=default_month)

if st.button("翌月の勤務表を生成する"):

    if "parsed_staff" not in st.session_state:
        st.warning("先に『既存勤務表を解析する』ボタンを押してください。")
        st.stop()

    with st.spinner("Gemini が勤務表を生成中…"):

        staff_json = st.session_state["parsed_staff"]
        codes_json = st.session_state["parsed_codes"]
        days = 30

        # 新人・パートの既存勤務を固定
        fixed_staff_schedule = {
            s["name"]: s["schedule"]
            for s in staff_json
            if s.get("role") in ["新入職員", "パート"]
        }

        # 生成対象（新人・パート以外）
        generate_target_staff = [
            s for s in staff_json
            if s.get("role") not in ["新入職員", "パート"]
        ]

        # 安全な % 方式でプロンプト生成
        generate_prompt = """
JSONのみ返してください。

以下の仕様で勤務表を生成してください。

{
  "month": "%s",
  "days": %d,
  "staff": %s
}

【勤務記号一覧】
%s

【勤務を生成する対象（新人・パート以外）】
%s

重要:
- 新入職員とパートの schedule は絶対に変更しない
- 生成対象は generate_target_staff のみ
- JSON以外の文章は禁止
- schedule は 1〜%d まで全て埋める
""" % (
            month,
            days,
            json.dumps(staff_json, ensure_ascii=False),
            json.dumps(codes_json, ensure_ascii=False),
            json.dumps(generate_target_staff, ensure_ascii=False),
            days
        )

        raw_output = gemini_generate(generate_prompt)
        st.write("### Gemini 生出力（生成）")
        st.text(raw_output)

        try:
            json_text = re.search(r'\{[\s\S]*\}', raw_output).group(0)
            generated_schedule = json.loads(json_text)
        except Exception as e:
            st.error(f"JSON解析に失敗しました: {e}")
            st.text(raw_output)
            st.stop()

        # 新人・パートの schedule を上書き（固定）
        for s in generated_schedule["staff"]:
            name = s["name"]
            if name in fixed_staff_schedule:
                s["schedule"] = fixed_staff_schedule[name]

        st.success("翌月の勤務表が生成されました！")
        st.json(generated_schedule)

        st.session_state["generated_schedule"] = generated_schedule


# ============================================================
# ③ 生成勤務表の手動編集UI
# ============================================================
if "generated_schedule" in st.session_state:
    st.write("## ③ 生成勤務表の手動編集")

    edited = {}

    for staff in st.session_state["generated_schedule"]["staff"]:
        st.write(f"### {staff['name']}（{staff['role']}）")

        schedule = staff["schedule"]
        new_schedule = {}

        cols = st.columns(10)
        for i in range(1, 31):
            col = cols[(i-1) % 10]
            new_schedule[str(i)] = col.text_input(
                f"{i}日",
                value=schedule.get(str(i), ""),
                key=f"{staff['name']}_{i}"
            )

        edited[staff["name"]] = new_schedule

    if st.button("編集内容を反映する"):
        for staff in st.session_state["generated_schedule"]["staff"]:
            staff["schedule"] = edited[staff["name"]]

        st.success("編集内容を反映しました！")


# ============================================================
# ④ Excel に既存勤務表の書式を完全コピーして書き出す
# ============================================================
st.write("## ④ Excel に既存勤務表の書式を完全コピーして書き出す")

if "generated_schedule" in st.session_state:

    generated_schedule = st.session_state["generated_schedule"]
    staff_list = generated_schedule.get("staff", [])
    days = generated_schedule.get("days", 30)

    rows = []

    # ヘッダー
    header = ["職員名", "役職", "グループ"] + [str(d) for d in range(1, days + 1)]
    rows.append(header)

    # 行データ
    for staff in staff_list:
        row = [
            staff.get("name", ""),
            staff.get("role", ""),
            staff.get("group", "")
        ] + [
            staff.get("schedule", {}).get(str(d), "")
            for d in range(1, days + 1)
        ]
        rows.append(row)

    df_out = pd.DataFrame(rows[1:], columns=rows[0])
    st.write("### 生成勤務表（テーブル表示）")
    st.dataframe(df_out)

    # 書式コピー
    wb = apply_format_to_generated_sheet(
        rows,
        format_map,
        col_widths,
        row_heights,
        merged_cells
    )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    st.download_button(
        label="既存勤務表の完全コピーをExcelでダウンロード",
        data=buffer,
        file_name=f"勤務表_{generated_schedule.get('month', 'unknown')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

else:
    st.info("まだ勤務表が生成されていません。『既存勤務表を解析する』『翌月の勤務表を生成する』を先に実行してください。")
